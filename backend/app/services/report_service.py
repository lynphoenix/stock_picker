# -*- coding: utf-8 -*-
"""
报表服务层
"""
import json
from pathlib import Path
from typing import Optional
import sys
from datetime import datetime

root_dir = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(root_dir))


class ReportService:
    """报表服务"""

    def __init__(self):
        self.tasks_dir = Path(root_dir) / "data" / "backtest_tasks"
        self.reports_dir = Path(root_dir) / "data" / "reports"
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    def generate_excel(self, task_id: str) -> Path:
        """
        生成Excel报表

        Args:
            task_id: 回测任务ID

        Returns:
            Excel文件路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # 加载回测结果
        result = self._load_result(task_id)
        if not result:
            raise FileNotFoundError(f"任务 {task_id} 的结果不存在")

        # 生成Excel文件
        excel_file = self.reports_dir / f"backtest_report_{task_id}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "概览"

        # 样式定义
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ===== 概览Sheet =====
        ws['A1'] = "回测报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        row = 3
        # 基本信息
        ws[f'A{row}'] = "基本信息"
        ws[f'A{row}'].font = title_font
        row += 1

        info_keys = [
            ("任务ID", "task_id"),
            ("策略", "strategy"),
            ("回测区间", lambda r: f"{r.get('start_date', '')} ~ {r.get('end_date', '')}"),
            ("初始资金", lambda r: f"¥{r.get('initial_capital', 0):,.2f}"),
            ("最终资金", lambda r: f"¥{r.get('final_capital', 0):,.2f}"),
        ]

        for label, key in info_keys:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = key(result) if callable(key) else result.get(key, '')
            row += 1

        row += 1
        # 绩效指标
        ws[f'A{row}'] = "绩效指标"
        ws[f'A{row}'].font = title_font
        row += 1

        metrics = [
            ("总收益率", lambda r: f"{r.get('total_return', 0):.2f}%"),
            ("年化收益率", lambda r: f"{r.get('annual_return', 0):.2f}%"),
            ("最大回撤", lambda r: f"{r.get('max_drawdown', 0):.2f}%"),
            ("夏普比率", lambda r: f"{r.get('sharpe_ratio', 0):.2f}"),
            ("胜率", lambda r: f"{r.get('win_rate', 0)*100:.1f}%"),
            ("交易次数", lambda r: r.get('total_trades', 0)),
            ("测试股票数", lambda r: r.get('stocks_tested', 0)),
        ]

        for label, func in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = func(result)
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25

        # ===== 交易明细Sheet =====
        if result.get('trades'):
            ws2 = wb.create_sheet("交易明细")
            headers = ["股票代码", "买入日期", "卖出日期", "买入价", "卖出价", "收益率", "持有天数"]
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            for row_idx, trade in enumerate(result['trades'], 2):
                ws2.cell(row_idx, 1, trade.get('code', ''))
                ws2.cell(row_idx, 2, trade.get('buy_date', ''))
                ws2.cell(row_idx, 3, trade.get('sell_date', ''))
                ws2.cell(row_idx, 4, trade.get('buy_price', 0))
                ws2.cell(row_idx, 5, trade.get('sell_price', 0))
                ws2.cell(row_idx, 6, f"{trade.get('return', 0)*100:.2f}%")
                ws2.cell(row_idx, 7, trade.get('hold_days', 0))

            for col in range(1, 8):
                ws2.column_dimensions[get_column_letter(col)].width = 12

        wb.save(excel_file)
        return excel_file

    def generate_pdf(self, task_id: str) -> Path:
        """
        生成PDF报表

        Args:
            task_id: 回测任务ID

        Returns:
            PDF文件路径
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image

        # 加载回测结果
        result = self._load_result(task_id)
        if not result:
            raise FileNotFoundError(f"任务 {task_id} 的结果不存在")

        # 生成PDF文件
        pdf_file = self.reports_dir / f"backtest_report_{task_id}.pdf"

        doc = SimpleDocTemplate(str(pdf_file), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # 标题样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
        )

        # 副标题样式
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
        )

        # ===== 标题 =====
        story.append(Paragraph("回测报告", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # ===== 基本信息 =====
        story.append(Paragraph("基本信息", subtitle_style))

        info_data = [
            ["项目", "内容"],
            ["任务ID", result.get('task_id', '')],
            ["策略", result.get('strategy', '')],
            ["回测区间", f"{result.get('start_date', '')} ~ {result.get('end_date', '')}"],
            ["初始资金", f"¥{result.get('initial_capital', 0):,.2f}"],
            ["最终资金", f"¥{result.get('final_capital', 0):,.2f}"],
        ]

        info_table = Table(info_data, colWidths=[2 * inch, 3 * inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))

        # ===== 绩效指标 =====
        story.append(Paragraph("绩效指标", subtitle_style))

        metrics_data = [
            ["指标", "数值"],
            ["总收益率", f"{result.get('total_return', 0):.2f}%"],
            ["年化收益率", f"{result.get('annual_return', 0):.2f}%"],
            ["最大回撤", f"{result.get('max_drawdown', 0):.2f}%"],
            ["夏普比率", f"{result.get('sharpe_ratio', 0):.2f}"],
            ["胜率", f"{result.get('win_rate', 0)*100:.1f}%"],
            ["交易次数", str(result.get('total_trades', 0))],
            ["测试股票数", str(result.get('stocks_tested', 0))],
        ]

        metrics_table = Table(metrics_data, colWidths=[2 * inch, 3 * inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(metrics_table)

        # ===== 交易明细 =====
        if result.get('trades'):
            story.append(Spacer(1, 0.3 * inch))
            story.append(Paragraph("交易明细", subtitle_style))

            trade_headers = [["股票代码", "买入日期", "卖出日期", "买入价", "卖出价", "收益率"]]
            for trade in result['trades'][:20]:  # 最多显示20条
                trade_headers.append([
                    trade.get('code', ''),
                    trade.get('buy_date', ''),
                    trade.get('sell_date', ''),
                    f"{trade.get('buy_price', 0):.2f}",
                    f"{trade.get('sell_price', 0):.2f}",
                    f"{trade.get('return', 0)*100:.2f}%"
                ])

            trade_table = Table(trade_headers)
            trade_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(trade_table)

        # ===== 页脚 =====
        story.append(Spacer(1, 0.5 * inch))
        story.append(Paragraph(
            f"<i>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>",
            styles['Normal']
        ))

        doc.build(story)
        return pdf_file

    def _load_result(self, task_id: str) -> Optional[dict]:
        """加载回测结果"""
        result_file = self.tasks_dir / f"{task_id}_result.json"
        if not result_file.exists():
            return None

        with open(result_file, 'r', encoding='utf-8') as f:
            return json.load(f)
